import requests
import openpyxl


crazy_shit = {}

def load_workbook():
    workbook = openpyxl.load_workbook('/Users/kierangarigan/Downloads/ScoutingSheetMaster-3.xlsm') #change this to the path of your scouting data
    worksheet = workbook.active
    start_row = 4
    end_row = 305
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1, max_col=worksheet.max_column))]
    data_dict = [dict(zip(headers, [cell.value for cell in row])) for row in worksheet.iter_rows(min_row=start_row, max_row=end_row)]

    for y in range(1, 57): #change 57 to number of matches + 1
        matchData = api(y)
        if not matchData:
            print(f"Failed to fetch data for match {y}")
            continue

        var = dict(matchData)
        api_scores = {'Blue': {'Auto': 0, 'TeleOP': 0}, 'Red': {'Auto': 0, 'TeleOP': 0}}

        for score in var['MatchScores']:
            for z in range(2):
                alliance_color = score['alliances'][z]['alliance']
                api_scores[alliance_color]['Auto'] = score['alliances'][z]['autoAmpNoteCount'] + score['alliances'][z]['autoSpeakerNoteCount']
                api_scores[alliance_color]['TeleOP'] = score['alliances'][z]['teleopAmpNoteCount'] + score['alliances'][z]['teleopSpeakerNoteCount'] + score['alliances'][z]['teleopSpeakerNoteAmplifiedCount']

        teleScoreRed, autoScoreRed, teleScoreBlue, autoScoreBlue = 0, 0, 0, 0

        for data in data_dict:
            if data['matchNum'] == y:
                if "b" in data["role"]: #You may need to change the strings depending on your sheet
                    teleScoreBlue += data['scoredInTeleop']
                    autoScoreBlue += data['scoredInAuto']
                elif "r" in data["role"]:
                    teleScoreRed += data['scoredInTeleop']
                    autoScoreRed += data['scoredInAuto']
        print("   ")
        print("   ")
        print("   ")
        for color in ['Blue', 'Red']:
            print("   ")
            print(f"{color} comparison for match {y}:")
            sheet_auto_score = autoScoreBlue if color == 'Blue' else autoScoreRed
            sheet_tele_score = teleScoreBlue if color == 'Blue' else teleScoreRed
            api_auto_score = api_scores[color]['Auto']
            api_tele_score = api_scores[color]['TeleOP']

            compare_scores(sheet_auto_score, api_auto_score, f"{color} Auto", y)
            compare_scores(sheet_tele_score, api_tele_score, f"{color} TeleOP", y)
        if y == 56:
            print("   ")
            print("   ")
            print("   ")
            print("   ")
            print("   ")
            print("   ")
            print("   ")
            print("Here are the > 5 note discrepancies:")
            print(crazy_shit)


def compare_scores(sheet_score, api_score, score_type, match_num):
    if sheet_score > api_score:
        print(f"Sheet {score_type} Score is higher than API by {sheet_score - api_score}")
        if sheet_score - api_score > 5:
            print("Large deviation in score noted. Discrepancy has been recorded")
            crazy_shit[(f'{score_type} Round {match_num}')] = (sheet_score - api_score)
    elif sheet_score < api_score:
        print(f"Sheet {score_type} Score is lower than API by {api_score - sheet_score}")
        if api_score - sheet_score > 5:
            print("Large deviation in score noted. Discrepancy has been recorded")
            crazy_shit[(f'{score_type} Round {match_num}')] = (api_score - sheet_score) 
    else:
        print(f"Sheet {score_type} Score is equal to API")
def api(match):
    url = f'https://frc-api.firstinspires.org/v3.0/2024/scores/WAYAK/Qualification?matchNumber={match}'
    api_key = 'YOUR FIRST API KEY GOES HERE'
    headers = {'Authorization': f'Basic {api_key}'}
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        print(f'Failed to get data: {response.status_code}')
        print(response.text)
        return None

load_workbook()
